import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _header_style():
    return {
        "fill": PatternFill("solid", fgColor="1a1a2e"),
        "font": Font(bold=True, color="FFFFFF"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def _label_style():
    return {
        "font": Font(bold=True, color="444444"),
        "fill": PatternFill("solid", fgColor="f5f5f5"),
        "alignment": Alignment(vertical="top", wrap_text=True),
    }


def _thin_border():
    thin = Side(style="thin", color="dddddd")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_row(ws, row_idx: int, values: list, style: dict | None = None):
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = _thin_border()
        if style:
            for attr, s in style.items():
                setattr(cell, attr, s)


def _write_title(ws, row_idx: int, title: str) -> int:
    cell = ws.cell(row=row_idx, column=1, value=title)
    cell.font = Font(bold=True, italic=True, color="333333")
    cell.border = _thin_border()
    return row_idx + 1


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def export_excel(page_results: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for page_num in sorted(page_results.keys()):
        page = page_results[page_num]
        sheet_name = f"Page {page_num}"[:31]
        ws = wb.create_sheet(title=sheet_name)
        row = 1
        wrote_anything = False

        for section in page.get("sections", []):
            stype = section.get("type")
            title = section.get("title")

            if stype == "table":
                columns = section.get("columns") or []
                rows = section.get("rows") or []
                if not columns:
                    continue
                if title:
                    row = _write_title(ws, row, title)
                _write_row(ws, row, columns, _header_style())
                row += 1
                for data_row in rows:
                    _write_row(ws, row, data_row)
                    row += 1
                row += 2
                wrote_anything = True

            elif stype == "key_value":
                pairs = section.get("pairs") or []
                if not pairs:
                    continue
                if title:
                    row = _write_title(ws, row, title)
                _write_row(ws, row, ["Key", "Value"], _header_style())
                row += 1
                for pair in pairs:
                    _write_row(ws, row, [pair.get("key", ""), pair.get("value", "")], _label_style())
                    row += 1
                row += 2
                wrote_anything = True

            elif stype == "qa_pair":
                items = section.get("items") or []
                if not items:
                    continue
                if title:
                    row = _write_title(ws, row, title)
                _write_row(ws, row, ["Question", "Answer"], _header_style())
                row += 1
                for item in items:
                    _write_row(ws, row, [item.get("question", ""), item.get("answer", "")])
                    row += 1
                row += 2
                wrote_anything = True

            elif stype == "paragraph":
                text = section.get("text", "")
                if title:
                    row = _write_title(ws, row, title)
                cell = ws.cell(row=row, column=1, value=text)
                cell.alignment = Alignment(wrap_text=True)
                cell.border = _thin_border()
                row += 3
                wrote_anything = True

        if not wrote_anything:
            ws.cell(row=1, column=1, value="No structured data extracted.")

        _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
