import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _style_header_row(ws, row_idx: int, num_cols: int):
    fill = PatternFill("solid", fgColor="1a1a2e")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_label_cell(cell):
    cell.font = Font(bold=True, color="444444")
    cell.fill = PatternFill("solid", fgColor="f5f5f5")
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _thin_border():
    thin = Side(style="thin", color="dddddd")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_excel(page_results: dict) -> bytes:
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for page_num in sorted(page_results.keys()):
            page = page_results[page_num]
            sheet_name = f"Page {page_num}"[:31]
            row_offset = 0

            # Collect dataframes to write per section
            wrote_anything = False

            for section in page.get("sections", []):
                stype = section.get("type")
                title = section.get("title")

                if stype == "table":
                    columns = section.get("columns") or []
                    rows = section.get("rows") or []
                    if not columns:
                        continue
                    df = pd.DataFrame(rows, columns=columns)

                    if title:
                        title_df = pd.DataFrame([[title]])
                        title_df.to_excel(writer, sheet_name=sheet_name,
                                          startrow=row_offset, index=False, header=False)
                        row_offset += 1

                    df.to_excel(writer, sheet_name=sheet_name,
                                startrow=row_offset, index=False)
                    row_offset += len(df) + 3
                    wrote_anything = True

                elif stype == "key_value":
                    pairs = section.get("pairs") or []
                    if not pairs:
                        continue
                    df = pd.DataFrame(pairs)

                    if title:
                        title_df = pd.DataFrame([[title]])
                        title_df.to_excel(writer, sheet_name=sheet_name,
                                          startrow=row_offset, index=False, header=False)
                        row_offset += 1

                    df.to_excel(writer, sheet_name=sheet_name,
                                startrow=row_offset, index=False)
                    row_offset += len(df) + 3
                    wrote_anything = True

                elif stype in ("qa_pair", "paragraph"):
                    # Write as a simple two-column or single-column text block
                    if stype == "qa_pair":
                        items = section.get("items") or []
                        rows_data = [
                            [item.get("question", ""), item.get("answer", "")]
                            for item in items
                        ]
                        df = pd.DataFrame(rows_data, columns=["Question", "Answer"])
                    else:
                        df = pd.DataFrame([[section.get("text", "")]], columns=["Text"])

                    if title:
                        title_df = pd.DataFrame([[title]])
                        title_df.to_excel(writer, sheet_name=sheet_name,
                                          startrow=row_offset, index=False, header=False)
                        row_offset += 1

                    df.to_excel(writer, sheet_name=sheet_name,
                                startrow=row_offset, index=False)
                    row_offset += len(df) + 3
                    wrote_anything = True

            if not wrote_anything:
                pd.DataFrame([["No structured data extracted."]]).to_excel(
                    writer, sheet_name=sheet_name, index=False, header=False
                )

        # Apply basic styling via openpyxl after writing
        wb = writer.book
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = _thin_border()
                    if cell.row == 1:
                        _style_label_cell(cell)
            # Auto-fit column widths (approximate)
            for col in ws.columns:
                max_len = max(
                    (len(str(c.value)) if c.value else 0 for c in col), default=0
                )
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    return buf.getvalue()
